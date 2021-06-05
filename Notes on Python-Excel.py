# Python Tutorial. Refer to page 290 of "Automate the Boring Stuff With Python"


# 1. GETTING STARTED
# To work with Excel in Python, you need to import the openpyxl module
# Go to File >> Settings >> Project >> Project Interpreter >> Click on the "+"
# Search for openpyxl, then install it. Then enter "import openpyxl" in the script.

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# Load the workbook you want to work with into Python using the ".load_workbook" method.
# Note that this workbook must be in the same directory as main.py (i.e. within the ExcelTutorial
# folder)
# Unlike in MATLAB, wb is the workbook data type (remember string, list, tuples = data types)
# Workbook data type: openpyxl.workbook.workbook.Workbook
wb = openpyxl.load_workbook('example.xlsx')

# Then from the workbook you need to load the sheet of interest by its name
sheet = wb['Sheet1']


# 2. ACCESSING A VALUE WITHIN A CELL
# The code below will print the value in cell A1, namely 4/5/2015 1:34:02 PM
cellA1 = sheet['A1'].value
print(cellA1)

# 3. A CELL'S ATTRIBUTES
# List below gives the row (1), the column (2 = B), the coordinate (B1) and the value within B1
cellUnknown = sheet['B1']
cellUnknownAttributes = [cellUnknown.row, cellUnknown.column, cellUnknown.coordinate, cellUnknown.value]
print(cellUnknownAttributes)

# 4. THE .CELL METHOD
# Cell in row 1, column 2 (B1) can be called using the .cell method
sheet.cell(row=1, column=2)

# Prints the values within cells of row = 1 3 5 7 (not including 8) + column 2
for i in range(1,8,2):
    print(i, sheet.cell(row=i, column=2).value)

# Side note: printing strings only, you must combine the strings using '+'. But you can't do that
# in the case above since i is an int.
# Combining string and int requires a comma instead.
print("Row of unknown cell is " + str(cellUnknown.row))

# 5. DETERMINING THE NUMBER OF ROWS AND COLUMNS
numberOfRows = sheet.max_row
numberOfColumns = sheet.max_column
print([numberOfRows, numberOfColumns])

# 6. CONVERTING COLUMN LETTERS TO #'S AND VICE VERSA
# get_column_letter(int): gives you the letter corresponding to the int'th column
# The code below should print AA, the 27th column
print(get_column_letter(27))

# column_index_from_string('str'): you pass the column letter 'str' and it returns the
# corresponding column number
# The code below should print 5 since the fifth column is associated with 'E'
print(column_index_from_string('E'))

# 7. ACCESSING ROWS AND COLUMNS
# The example.xlsx contains 7 rows + 3 columns
# How do we iterate through all the cells in, say, column B?
# 1st, we have to convert the entire sheet into a tuple
sheetAsTuple = tuple(sheet.rows)

# This is a tuple consisting of 7 sub-tuples
# Each sub-tuple consists of all the values in a given row
# ((<Sheet1.A1>, <Sheet1.B1>, <Sheet1.C1>, ...),
#  (<Sheet1.A2>, <Sheet1.B2>, <Sheet1.C2>, ...),
#  (<Sheet1.A3>, <Sheet1.B3>, <Sheet1.C3>,...), ...)

for index in range(0,sheet.max_row):
    print(sheetAsTuple[index][1])

# Please remember: tuples, lists, all start with index 0!!!
# Also: range(0,3) reads index = 0, 1, 2 (but not 3)!

# Prints cells A1:B2
for rowIndex in range(0,2):
    for columnIndex in range(0,2):
        print(sheetAsTuple[rowIndex][columnIndex])
















