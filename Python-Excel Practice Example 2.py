# Writing to Excel File in Python Tutorial
# Automate the Boring Stuff With Python, page 279

# This script takes the file 'produceSales.xlsx' and updates the prices of garlic, celery and lemon
# Updated cost per pound:
# Celery    1.19
# Garlic    3.07
# Lemon     1.27
# Save the updated spreadsheet to a new file called 'produceSales_rev1.xlsx'

import openpyxl

# import the workbook and load the worksheet
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# Create a dictionary consisting of the updated prices
updatedPrices = {'Celery': 1.19,
                 'Garlic': 3.07,
                 'Lemon': 1.27}

# Loop through each row and if the produce associated with that row matches one of the keys in the
# dictionary updatedPrices, update its price!
for rowIndex in range(2,sheet.max_row+1):
    produce = sheet['A'+str(rowIndex)].value

    # Note: the in method is used to determine if produce is an index in the dictionary updatedPrices
    # It can also be used to determine if a given index is in a list or a tuple too!
    # e.g. if index in updatedPrices.values():
    # The commented code above checks if index is within the list of values of the dictionary updatedPrices
    if produce in updatedPrices:
        # Alternative to sheet['B'+str(rowIndex)]...
        # sheet.cell(row=rowIndex, column=2)
        # For excel, the first row/column has an index of 1, unlike in dictionaries/tuples/lists where
        # the first index is 0.
        sheet['B'+str(rowIndex)] = updatedPrices[produce]


# Write the results to a copy of the file, not the original file!
wb.save('produceSales_rev1.xlsx')

