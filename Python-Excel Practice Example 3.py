# Implementing Python Code to Generate a Multiplication Table in Excel
# Automate the Boring Stuff With Python, page 291

# This script takes a number N as input and prints out an NxN multiplication table in an
# Excel workbook entitled 'multiplication_table.xlsx'

import openpyxl

# Import the Font() function
from openpyxl.styles import Font

# Initiate a new workbook and load the first worksheet
wb = openpyxl.load_workbook('multiplication_table.xlsx')
sheet = wb['Sheet1']


# Prompt the user for a value of N
print('Enter the value of N:')
N = int(input())

# Initiate the headings of the multiplication table (row 1 and column A)
for index in range(2,N+2):
    sheet.cell(row=1, column=index).value = index-1
    sheet.cell(row=index, column=1).value = index-1

    # Set the font of the headings to bold
    sheet.cell(row=1, column=index).font = Font(bold=True)
    sheet.cell(row=index, column=1).font = Font(bold=True)

# Populate the multiplication table
for columnIndex in range(2,N+2):
    for rowIndex in range(2,N+2):
        # Store the values within the headers
        verticalHeaderValue = sheet.cell(row=rowIndex, column=1).value
        horizontalHeaderValue = sheet.cell(row=1, column=columnIndex).value

        # Multiply both values together and store this in the cell of
        # row=rowIndex, column=columnIndex
        sheet.cell(row=rowIndex, column=columnIndex).value = verticalHeaderValue * horizontalHeaderValue

wb.save('multiplication_table.xlsx')









